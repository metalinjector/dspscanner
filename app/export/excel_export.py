"""Экспорт результатов в Excel (.xlsx) через openpyxl."""
from __future__ import annotations

import math
import os
import re
import tempfile
from pathlib import Path
from typing import Sequence

from app.config import ErrorEntry, ScanStats, SearchResult
from app.export.safety import spreadsheet_text

_FILENAME_CONTEXT_PREFIX = "[Имя файла]"
_CONTEXT_COLUMN_CHARS = 70
_CONTEXT_MIN_WIDTH = 40
_CONTEXT_MAX_WIDTH = 90


def export_to_excel(
    results: Sequence[SearchResult],
    stats: ScanStats,
    errors: Sequence[ErrorEntry],
    search_info: dict,
    file_path: str,
) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        return False

    target = Path(file_path)
    tmp_path: Path | None = None
    try:
        target.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        _build_results_sheet(
            workbook.active,
            results,
            Font,
            PatternFill,
            Alignment,
            get_column_letter,
            Table,
            TableStyleInfo,
            CellRichText,
            TextBlock,
            InlineFont,
        )
        workbook.active.title = "Результаты"
        _build_summary_sheet(workbook.create_sheet("Сводка"), stats, search_info, Font)
        _build_errors_sheet(
            workbook.create_sheet("Ошибки"), errors, Font, PatternFill,
            get_column_letter, Table, TableStyleInfo,
        )

        fd, tmp_name = tempfile.mkstemp(prefix=f".{target.stem}.", suffix=".xlsx", dir=target.parent)
        os.close(fd)
        tmp_path = Path(tmp_name)
        workbook.save(tmp_path)
        os.replace(tmp_path, target)
        return True
    except Exception:
        if tmp_path is not None:
            try:
                tmp_path.unlink(missing_ok=True)
            except OSError:
                pass
        return False


def _rich_context(text: str, matched_text: str, CellRichText, TextBlock, InlineFont):
    """Возвращает rich-text с красным выделением всех найденных фрагментов."""
    needle = (matched_text or "").strip()
    if not text or not needle:
        return text

    matches = list(re.finditer(re.escape(needle), text, flags=re.IGNORECASE))
    if not matches:
        return text

    red_font = InlineFont(color="FFFF0000", b=True)
    parts: list[object] = []
    cursor = 0
    for match in matches:
        if match.start() > cursor:
            parts.append(text[cursor:match.start()])
        parts.append(TextBlock(red_font, text[match.start():match.end()]))
        cursor = match.end()
    if cursor < len(text):
        parts.append(text[cursor:])
    return CellRichText(parts)


def _estimated_row_height(text: str, column_chars: int = _CONTEXT_COLUMN_CHARS) -> float:
    """Приближённый AutoFit высоты: openpyxl сам не умеет вычислять её."""
    if not text:
        return 18.0
    visual_lines = 0
    for line in text.splitlines() or [text]:
        visual_lines += max(1, math.ceil(len(line) / max(1, column_chars)))
    return min(180.0, max(18.0, 15.0 * visual_lines))



def _automatic_context_width(results: Sequence[SearchResult]) -> int:
    """Подбирает ширину столбца по данным, но не создаёт чрезмерно широкую таблицу."""
    longest = 0
    for result in results:
        text = spreadsheet_text(result.context)
        longest = max(longest, *(len(line) for line in text.splitlines() or [text]))
    if not results:
        return _CONTEXT_COLUMN_CHARS
    return min(_CONTEXT_MAX_WIDTH, max(_CONTEXT_MIN_WIDTH, longest + 2))

def _build_results_sheet(
    ws,
    results,
    Font,
    PatternFill,
    Alignment,
    get_column_letter,
    Table,
    TableStyleInfo,
    CellRichText,
    TextBlock,
    InlineFont,
):
    headers = ["№", "Имя файла", "Полный путь", "Тип", "Найденное слово", "Контекст", "Изменён"]
    ws.append(headers)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    filename_row_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for index in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=index)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    context_column_width = _automatic_context_width(results)

    for index, result in enumerate(results, start=1):
        row_index = index + 1
        context = spreadsheet_text(result.context)
        ws.append([
            index,
            spreadsheet_text(result.file_name),
            spreadsheet_text(result.full_path),
            spreadsheet_text(result.file_type),
            spreadsheet_text(result.word),
            context,
            spreadsheet_text(result.modified),
        ])

        # Выделяем именно фактически найденный текст, а не шаблон с '*'.
        matched_text = result.matched_text or result.word
        ws.cell(row=row_index, column=6).value = _rich_context(
            context, matched_text, CellRichText, TextBlock, InlineFont,
        )
        ws.cell(row=row_index, column=6).alignment = Alignment(
            wrap_text=True, vertical="top",
        )
        ws.row_dimensions[row_index].height = _estimated_row_height(context, context_column_width)

        # Совпадение в названии файла визуально отличается от совпадения в содержимом.
        if result.context.startswith(_FILENAME_CONTEXT_PREFIX):
            for column in range(1, len(headers) + 1):
                ws.cell(row=row_index, column=column).fill = filename_row_fill

        for column in range(1, len(headers) + 1):
            if column != 6:
                ws.cell(row=row_index, column=column).alignment = Alignment(vertical="top")

    # Контекст получает перенос строк, адаптивную ширину и практический AutoFit высоты.
    for index, width in enumerate([6, 28, 55, 8, 20, context_column_width, 18], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{max(1, len(results) + 1)}"

    if results:
        try:
            table = Table(displayName="ResultsTable", ref=f"A1:G{len(results) + 1}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(table)
        except Exception:
            pass


def _build_summary_sheet(ws, stats: ScanStats, search_info: dict, Font):
    ws["A1"] = "Отчёт DSP Scanner"
    ws["A1"].font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    rows = [
        ("Дата создания отчёта", spreadsheet_text(search_info.get("generated_at", ""))),
        ("Пути поиска", spreadsheet_text(", ".join(map(str, search_info.get("paths", []))))),
        ("Термины в содержимом", spreadsheet_text(", ".join(map(str, search_info.get("words", []))))),
        ("Термины в названиях файлов", spreadsheet_text(", ".join(map(str, search_info.get("filename_words", []))))),
        ("Метод чтения DOC", spreadsheet_text(search_info.get("doc_method", ""))),
        ("", ""),
        ("Обработано файлов", stats.processed),
        ("  DOCX", stats.docx_count),
        ("  DOC", stats.doc_count),
        ("  TXT", stats.txt_count),
        ("  PDF", stats.pdf_count),
        ("Найдено совпадений", stats.found),
        ("Пропущено", stats.skipped),
        ("Временных файлов пропущено", stats.temp_files),
        ("Заблокированных файлов", stats.locked),
        ("Повреждённых файлов", stats.corrupted),
        ("Ошибок", stats.errors),
    ]
    for row_index, (label, value) in enumerate(rows, start=3):
        ws.cell(row=row_index, column=1, value=label).font = label_font
        ws.cell(row=row_index, column=2, value=value)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 60


def _build_errors_sheet(ws, errors, Font, PatternFill, get_column_letter, Table, TableStyleInfo):
    headers = ["Время", "Уровень", "Файл", "Сообщение"]
    ws.append(headers)
    fill = PatternFill(start_color="7A2222", end_color="7A2222", fill_type="solid")
    font = Font(bold=True, color="FFFFFF")
    for index in range(1, len(headers) + 1):
        ws.cell(row=1, column=index).font = font
        ws.cell(row=1, column=index).fill = fill
    for error in errors:
        ws.append([
            spreadsheet_text(error.timestamp),
            spreadsheet_text(error.level),
            spreadsheet_text(error.file_path),
            spreadsheet_text(error.message),
        ])
    for index, width in enumerate([20, 10, 55, 70], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A2"
    if errors:
        try:
            table = Table(displayName="ErrorsTable", ref=f"A1:D{len(errors) + 1}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium3", showRowStripes=True)
            ws.add_table(table)
        except Exception:
            pass
